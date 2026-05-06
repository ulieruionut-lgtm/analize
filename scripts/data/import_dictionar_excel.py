"""
Importă dictionarul de analize din Excel (dictionar_analize_300_1200_alias.xlsx)
în baza de date: analiza_standard + analiza_alias.

Structura Excel: sheet "Analyte_Alias", coloane: analyte (denumire standard), alias

Rulează: python import_dictionar_excel.py
Sau: IMPORT_DICTIONAR_EXCEL.bat

Pentru Railway (producție): copiază DATABASE_URL din Railway Variables în .env,
apoi rulează scriptul - va actualiza baza de date de producție.
"""
import re
import sys
import unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

# Încarcă .env înainte de orice import din backend (pentru DATABASE_URL)
try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / ".env")
except ImportError:
    pass


def _to_cod(s: str) -> str:
    """Generează cod unic din denumire: majuscule, fără diacritice, spații -> _"""
    if not s or not str(s).strip():
        return "ANALIZA"
    s = str(s).upper().strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^\w\s]", "", s)
    s = re.sub(r"\s+", "_", s).strip("_")
    return s[:64] if s else "ANALIZA"


def main():
    try:
        import openpyxl
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    from backend.database import (
        _get_or_create_analiza_standard,
        get_cursor,
        _use_sqlite,
    )
    from backend.normalizer import invalideaza_cache

    def _find_analiza_by_denumire(denumire: str):
        """Cauta analiza_standard dupa denumire_standard (match normalizat)."""
        if not denumire or not denumire.strip():
            return None
        dn = str(denumire).strip().lower()
        with get_cursor(commit=False) as cur:
            if _use_sqlite():
                cur.execute(
                    "SELECT id FROM analiza_standard WHERE LOWER(TRIM(denumire_standard)) = ?",
                    (dn,),
                )
            else:
                cur.execute(
                    "SELECT id FROM analiza_standard WHERE LOWER(TRIM(denumire_standard)) = LOWER(TRIM(%s))",
                    (denumire,),
                )
            row = cur.fetchone()
            if row:
                return row[0]
        return None

    xlsx_path = ROOT / "dictionar_analize_300_1200_alias.xlsx"
    if not xlsx_path.exists():
        print("EROARE: Fisier negasit:", xlsx_path)
        return 1

    print("Golire alias-uri existente (vor fi reinportate din Excel)...")
    with get_cursor() as cur:
        if _use_sqlite():
            cur.execute("DELETE FROM analiza_alias")
        else:
            cur.execute("DELETE FROM analiza_alias")

    print("Citire Excel...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = wb["Analyte_Alias"] if "Analyte_Alias" in wb.sheetnames else wb.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))  # skip header
    wb.close()

    # Mapare analyte -> analiza_standard_id
    analyte_to_id = {}
    stats = {"analize_noi": 0, "aliasuri_adaugate": 0, "aliasuri_skip": 0, "erori": 0}

    print("Import analize standard si aliasuri...")
    for i, (analyte, alias) in enumerate(rows):
        try:
            analyte = (analyte or "").strip()
            alias = (alias or "").strip()
            if not analyte or not alias:
                continue
            if analyte not in analyte_to_id:
                aid = _find_analiza_by_denumire(analyte)
                if aid is None:
                    cod = _to_cod(analyte)
                    aid = _get_or_create_analiza_standard(cod, analyte)
                    stats["analize_noi"] += 1
                analyte_to_id[analyte] = aid
            aid = analyte_to_id[analyte]
            with get_cursor(commit=True) as cur:
                if _use_sqlite():
                    cur.execute(
                        "INSERT OR IGNORE INTO analiza_alias (analiza_standard_id, alias) VALUES (?, ?)",
                        (aid, alias),
                    )
                else:
                    cur.execute(
                        """INSERT INTO analiza_alias (analiza_standard_id, alias) VALUES (%s, %s)
                           ON CONFLICT (alias) DO NOTHING""",
                        (aid, alias),
                    )
                stats["aliasuri_adaugate"] += 1
        except Exception as e:
            stats["erori"] += 1
            if stats["erori"] <= 5:
                print(f"  Eroare rand {i+2}: {e}")

    invalideaza_cache()
    print()
    print("=== IMPORT FINALIZAT ===")
    print(f"  Analize standard (unic): {len(analyte_to_id)}")
    print(f"  Aliasuri procesate: {len(rows)}")
    print(f"  Erori: {stats['erori']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
