"""Importă dictionarul de analize din Excel în baza de date."""
import re
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, Dict, List, Tuple, Union


def _to_cod(s: str) -> str:
    """Generează cod unic din denumire: majuscule, fără diacritice, spații -> _"""
    if not s or not str(s).strip():
        return "ANALIZA"
    s = str(s).upper().strip()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^\w\s]", "", s)
    s = re.sub(r"\s+", "_", s).strip("_")
    return s[:64] if s else "ANALIZA"


def import_dictionar_excel(
    source: Union[str, Path, bytes, BinaryIO],
    progress_callback=None,
) -> Dict:
    """
    Importă analize standard + aliasuri din Excel.
    source: cale fișier, bytes, sau file-like.
    Returnează: {ok, analize_unic, aliasuri_procesate, erori, mesaj}
    """
    from backend.database import (
        _get_or_create_analiza_standard,
        get_cursor,
        _use_sqlite,
    )
    from backend.normalizer import invalideaza_cache

    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "mesaj": "Lipsește openpyxl. Instalează: pip install openpyxl"}

    if isinstance(source, (str, Path)):
        wb = openpyxl.load_workbook(str(source), read_only=True, data_only=True)
    elif isinstance(source, bytes):
        wb = openpyxl.load_workbook(BytesIO(source), read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)

    sheet = wb["Analyte_Alias"] if "Analyte_Alias" in wb.sheetnames else wb.active
    rows: List[Tuple] = list(sheet.iter_rows(min_row=2, values_only=True))
    wb.close()

    with get_cursor() as cur:
        if _use_sqlite():
            cur.execute("DELETE FROM analiza_alias")
        else:
            cur.execute("DELETE FROM analiza_alias")

    with get_cursor(commit=False) as cur:
        cur.execute("SELECT id, denumire_standard FROM analiza_standard")
        denumire_to_id = {}
        for row in cur.fetchall():
            rid, den_val = row[0], (row[1] or "")
            den = str(den_val).strip().lower()
            if den:
                denumire_to_id[den] = rid

    analyte_to_id = {}
    erori = 0
    total = len(rows)
    batch = []
    BATCH_SIZE = 50
    if progress_callback and total > 0:
        progress_callback(0, total)
    for i, row in enumerate(rows):
        if progress_callback and (i + 1) % 50 == 0:
            progress_callback(i + 1, total)
        if not row or len(row) < 2:
            continue
        analyte, alias = (row[0] or "").strip(), (row[1] or "").strip()
        if not analyte or not alias:
            continue
        try:
            if analyte not in analyte_to_id:
                dn = analyte.strip().lower()
                aid = denumire_to_id.get(dn)
                if aid is None:
                    cod = _to_cod(analyte)
                    aid = _get_or_create_analiza_standard(cod, analyte)
                    denumire_to_id[dn] = aid
                analyte_to_id[analyte] = aid
            aid = analyte_to_id[analyte]
            batch.append((aid, alias))
        except Exception:
            erori += 1
            continue
        if len(batch) >= BATCH_SIZE:
            try:
                with get_cursor() as cur:
                    if _use_sqlite():
                        cur.executemany(
                            "INSERT OR IGNORE INTO analiza_alias (analiza_standard_id, alias) VALUES (?, ?)",
                            batch,
                        )
                    else:
                        cur.executemany(
                            """INSERT INTO analiza_alias (analiza_standard_id, alias) VALUES (%s, %s)
                               ON CONFLICT (alias) DO NOTHING""",
                            batch,
                        )
                batch = []
            except Exception:
                erori += len(batch)
                batch = []
    if batch:
        try:
            with get_cursor() as cur:
                if _use_sqlite():
                    cur.executemany(
                        "INSERT OR IGNORE INTO analiza_alias (analiza_standard_id, alias) VALUES (?, ?)",
                        batch,
                    )
                else:
                    cur.executemany(
                        """INSERT INTO analiza_alias (analiza_standard_id, alias) VALUES (%s, %s)
                           ON CONFLICT (alias) DO NOTHING""",
                        batch,
                    )
        except Exception:
            erori += len(batch)

    if progress_callback and total > 0:
        progress_callback(total, total)

    invalideaza_cache()
    return {
        "ok": True,
        "analize_unic": len(analyte_to_id),
        "aliasuri_procesate": len(rows),
        "erori": erori,
        "mesaj": f"Import finalizat: {len(analyte_to_id)} analize, {len(rows)} aliasuri.",
    }
